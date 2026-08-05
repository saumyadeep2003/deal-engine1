"""Live Google Sheets mirror of the workbook, with the same two-way sync contract.

The nine required tabs (plus Provenance and Demo Cases) are pushed to a Google
Sheet partners can open and edit from anywhere. Before every push, the
Recommendation column is read back from the sheet and written to the DB —
**the human value wins** and the disagreement is logged to `partner_actions`,
exactly as with the local .xlsx.

Auth: a Google service account. Without credentials this module does NOT fail the
pipeline — it records `not_configured` in `sheet_sync` and the local workbook
remains the source of truth. Setup is documented in RUNBOOK.md §Google Sheets.

Env:
  GOOGLE_SERVICE_ACCOUNT_JSON  path to the service-account key file
  GSHEET_ID                    id of an existing sheet (preferred), or
  GSHEET_TITLE                 title to create/open by name (default 'Thirdbase Deal Pipeline')
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from engine import db  # noqa: E402
from engine.config import env  # noqa: E402

SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive.file"]

REC_COL_INDEX = 15          # 1-based column O on Pipeline-shaped tabs
EDITABLE_TABS = ("Pipeline", "Hot Deals", "Watchlist")


def credentials_path() -> Path | None:
    p = env("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not p:
        return None
    path = Path(p).expanduser()
    return path if path.exists() else None


def configured() -> bool:
    return credentials_path() is not None


def status() -> dict:
    last = db.q1("SELECT * FROM sheet_sync ORDER BY synced_at DESC LIMIT 1")
    return {"configured": configured(),
            "credentials": str(credentials_path()) if configured() else None,
            "reason": None if configured() else
                      "GOOGLE_SERVICE_ACCOUNT_JSON not set or file missing",
            "last_sync": dict(last) if last else None}


def _client():
    import gspread
    from google.oauth2.service_account import Credentials
    creds = Credentials.from_service_account_file(str(credentials_path()), scopes=SCOPES)
    return gspread.authorize(creds), creds


def _open_sheet(gc, creds):
    sheet_id = env("GSHEET_ID")
    if sheet_id:
        return gc.open_by_key(sheet_id)
    title = env("GSHEET_TITLE", "Thirdbase Deal Pipeline")
    try:
        return gc.open(title)
    except Exception:  # noqa: BLE001 — not found, create it
        sh = gc.create(title)
        share_to = env("GSHEET_SHARE_WITH")
        if share_to:
            for addr in share_to.split(","):
                try:
                    sh.share(addr.strip(), perm_type="user", role="writer")
                except Exception:  # noqa: BLE001
                    pass
        print(f"  created Google Sheet '{title}' — {sh.url}")
        print("  NOTE: a service-account sheet is not in your Drive UI until shared."
              " Set GSHEET_SHARE_WITH=you@gmail.com, or use GSHEET_ID of a sheet you"
              " already shared with the service-account email.")
        return sh


def pull_human_edits(sh, partner: str = "partner", verbose: bool = True) -> int:
    """Read Recommendation edits from the sheet into the DB. Human value wins."""
    n = 0
    seen: set[str] = set()
    for tab in EDITABLE_TABS:
        try:
            ws = sh.worksheet(tab)
        except Exception:  # noqa: BLE001 — tab not created yet
            continue
        rows = ws.get_all_values()
        for row in rows[1:]:
            if not row or not row[0] or row[0].startswith("Legend:"):
                continue
            name = row[0]
            if name in seen or len(row) < REC_COL_INDEX:
                continue
            seen.add(name)
            rec_val = (row[REC_COL_INDEX - 1] or "").strip()
            if rec_val not in ("Pass", "Watch", "Deep Dive"):
                continue
            comp = db.q1("SELECT id FROM companies WHERE name=? AND is_synthetic=0", (name,))
            if not comp:
                continue
            score = db.q1("SELECT id, recommendation, human_override, composite, features_json"
                          " FROM scores WHERE company_id=? ORDER BY scored_at DESC, id DESC"
                          " LIMIT 1", (comp["id"],))
            if not score:
                continue
            effective = score["human_override"] or score["recommendation"]
            if rec_val != effective:
                db.execute("UPDATE scores SET human_override=? WHERE id=?",
                           (rec_val, score["id"]))
                db.insert("partner_actions", {
                    "company_id": comp["id"], "partner": partner, "action": "override",
                    "score_at_time": score["composite"],
                    "features_at_time_json": score["features_json"],
                    "note": f"google sheet edit: {effective} -> {rec_val} ({tab})",
                    "created_at": db.now_iso()})
                n += 1
    if verbose and n:
        print(f"  google sheet: {n} human edit(s) pulled into DB (human value wins)")
    return n


def _tab_data(xlsx_path: Path) -> list[tuple[str, list[list]]]:
    """Read the generated workbook so the Sheet is a faithful mirror — one
    renderer, two destinations, no second source of truth."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        out.append((name, rows))
    return out


def sync(xlsx_path: Path | None = None, verbose: bool = True) -> dict:
    from engine.config import OUTPUT_DIR
    xlsx_path = xlsx_path or (OUTPUT_DIR / "deal_pipeline.xlsx")

    if not configured():
        rec = {"status": "not_configured", "tabs_written": 0, "edits_pulled": 0,
               "detail": "GOOGLE_SERVICE_ACCOUNT_JSON not set — local workbook"
                         " remains source of truth", "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  google sheets: skipped — {rec['detail']}")
        return rec
    if not xlsx_path.exists():
        rec = {"status": "error", "detail": f"{xlsx_path} not generated yet",
               "tabs_written": 0, "edits_pulled": 0, "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        return rec

    try:
        gc, creds = _client()
        sh = _open_sheet(gc, creds)
        edits = pull_human_edits(sh, verbose=verbose)   # read BEFORE writing
        tabs = _tab_data(xlsx_path)
        existing = {w.title for w in sh.worksheets()}
        for name, rows in tabs:
            n_rows = max(len(rows), 2)
            n_cols = max((len(r) for r in rows), default=1)
            if name in existing:
                ws = sh.worksheet(name)
                ws.clear()
                ws.resize(rows=n_rows, cols=n_cols)
            else:
                ws = sh.add_worksheet(title=name, rows=n_rows, cols=n_cols)
            if rows:
                ws.update(values=rows, range_name="A1")
                ws.freeze(rows=1)
                ws.format(f"A1:{chr(64 + min(n_cols, 26))}1",
                          {"textFormat": {"bold": True,
                                          "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                           "backgroundColor": {"red": 0.12, "green": 0.23, "blue": 0.34}})
        for stale in existing - {n for n, _ in tabs} - {"Sheet1"}:
            pass  # never delete partner-created tabs
        rec = {"status": "ok", "spreadsheet_id": sh.id, "spreadsheet_url": sh.url,
               "tabs_written": len(tabs), "edits_pulled": edits,
               "detail": None, "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  google sheets: {len(tabs)} tabs synced -> {sh.url}")
        return rec
    except Exception as exc:  # noqa: BLE001 — never kill the pipeline for a sync
        rec = {"status": "error", "tabs_written": 0, "edits_pulled": 0,
               "detail": f"{type(exc).__name__}: {exc}"[:400], "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  ! google sheets sync failed — {rec['detail']}")
        return rec


if __name__ == "__main__":
    if "--status" in sys.argv:
        print(json.dumps(status(), indent=2, default=str))
    else:
        print(json.dumps(sync(), indent=2, default=str))
