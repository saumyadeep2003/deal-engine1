"""Component 09 — Excel writer. Nine required tabs + Provenance + Demo Cases.

Two-way sync: BEFORE regenerating, human-editable columns are read from the
existing workbook and written back to the DB (partner_actions logs the
disagreement; the human value wins). Synthetic records are amber-highlighted
and confined to Demo Cases; every tab carries the legend.

Missing licence-gated fields render as '— (requires X)' — visibly incomplete
rather than invisibly wrong.
"""
from __future__ import annotations
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from engine import db, scoring  # noqa: E402
from engine.config import OUTPUT_DIR, thesis  # noqa: E402

WORKBOOK = OUTPUT_DIR / "deal_pipeline.xlsx"

PIPELINE_COLS = [
    "Company", "One-line description", "Sector", "Stage", "Last round size",
    "Last round date", "Valuation", "Lead investor", "Tier 1 count", "Headcount",
    "6-month growth", "Thesis score (cohort pctile)", "Theme tag",
    "Last signal date", "Recommendation", "Link to full brief",
    "Investor commentary summary",
]
# Note: the fund's brief enumerates these 17 columns for the Pipeline tab
# (its prose says 16 — the enumerated list is authoritative; see BUILD_LOG.md).

HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
HEADER_FONT = Font(color="FFFFFF", bold=True)
AMBER_FILL = PatternFill("solid", fgColor="FFBF00")
LEGEND = ("Legend: amber rows = SYNTHETIC demo records (confined to the Demo Cases tab). "
          "'— (requires X)' = field needs a licensed source; real tabs contain real data only.")
LEGEND_FONT = Font(italic=True, size=9, color="808080")


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")[:60]


def _fmt_money(v) -> str | None:
    if v is None:
        return None
    v = float(v)
    if v >= 1e9:
        return f"${v / 1e9:.2f}B"
    if v >= 1e6:
        return f"${v / 1e6:.1f}M"
    return f"${v:,.0f}"


def _sheet(wb: Workbook, title: str, headers: list[str]) -> "object":
    ws = wb.create_sheet(title)
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(42, len(h) + 8))
    return ws


def _legend(ws) -> None:
    """Legend on every tab — placed beside the header row so it never pollutes
    the data range (autofilter/two-way sync read contiguous rows)."""
    ws.oddHeader.center.text = LEGEND        # printed header
    c = ws.cell(row=1, column=ws.max_column + 2, value=LEGEND)
    c.font = LEGEND_FONT


# --------------------------------------------------------------------------
# Two-way sync: read human edits BEFORE regenerating. Human edits always win.
# --------------------------------------------------------------------------

def read_human_edits(partner: str = "partner") -> int:
    if not WORKBOOK.exists():
        return 0
    try:
        wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    except Exception:
        return 0
    n = 0
    seen: set[str] = set()   # first occurrence wins — Pipeline is the master edit surface
    for tab in ("Pipeline", "Hot Deals", "Watchlist"):
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = ws.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or not row[0] or str(row[0]).startswith("Legend:"):
                continue
            name, rec_cell = str(row[0]), row[14] if len(row) > 14 else None
            if name in seen:
                continue
            seen.add(name)
            if not rec_cell:
                continue
            rec_val = str(rec_cell).strip()
            if rec_val not in ("Pass", "Watch", "Deep Dive"):
                continue
            comp = db.q1("SELECT id FROM companies WHERE name=? AND is_synthetic=0", (name,))
            if not comp:
                continue
            score = db.q1("SELECT id, recommendation, human_override, composite, features_json"
                          " FROM scores WHERE company_id=? ORDER BY scored_at DESC, id DESC LIMIT 1",
                          (comp["id"],))
            if not score:
                continue
            effective = score["human_override"] or score["recommendation"]
            if rec_val != effective:
                db.execute("UPDATE scores SET human_override=? WHERE id=?", (rec_val, score["id"]))
                db.insert("partner_actions", {
                    "company_id": comp["id"], "partner": partner, "action": "override",
                    "score_at_time": score["composite"],
                    "features_at_time_json": score["features_json"],
                    "note": f"workbook edit: {effective} -> {rec_val} ({tab})",
                    "created_at": db.now_iso()})
                n += 1
    wb.close()
    if n:
        print(f"  two-way sync: {n} human edit(s) written back to DB (human value wins)")
    return n


# --------------------------------------------------------------------------
# Row assembly
# --------------------------------------------------------------------------

def _pipeline_row(c: dict) -> list:
    cid = c["id"]
    rnd = db.q1("SELECT fr.amount_usd, fr.valuation_usd, fr.announced_at, fr.stage, i.name lead"
                " FROM funding_rounds fr LEFT JOIN investors i ON fr.lead_investor_id=i.id"
                " WHERE fr.company_id=? ORDER BY fr.announced_at DESC LIMIT 1", (cid,))
    feats = json.loads(c["features_json"])["computed"] if c.get("features_json") else {}
    tier1 = feats.get("tier1_count", {}).get("value", 0)
    headcount = db.q1("SELECT value_json, unavailable_reason FROM enrichment_cache"
                      " WHERE company_id=? AND field='headcount'", (cid,))
    growth = db.q1("SELECT value_json, unavailable_reason FROM enrichment_cache"
                   " WHERE company_id=? AND field='headcount_growth_6m'", (cid,))

    def gated(row) -> str:
        if row and row["value_json"]:
            return json.loads(row["value_json"])
        if row and row["unavailable_reason"]:
            return f"— ({row['unavailable_reason']})"
        return "—"

    brief = db.q1("SELECT id FROM briefs WHERE company_id=? AND validated=1"
                  " ORDER BY generated_at DESC LIMIT 1", (cid,))
    brief_link = f"briefs/{_slug(c['name'])}.md" if brief else "—"
    comm = db.q1("SELECT COUNT(*) n FROM commentary WHERE company_id=?", (cid,))
    comm_sum = db.q1("SELECT quote FROM commentary WHERE company_id=? ORDER BY observed_at DESC"
                     " LIMIT 1", (cid,))
    lc = " (low-confidence cohort <20)" if c.get("cohort_low_confidence") else ""
    val = rnd["valuation_usd"] if rnd else None
    return [
        c["name"],
        (c["description"] or "")[:160] or "—",
        c.get("sub_sector") or c.get("sector") or "unclassified",
        (rnd["stage"] if rnd and rnd["stage"] else c.get("stage")) or "unknown",
        _fmt_money(rnd["amount_usd"]) if rnd and rnd["amount_usd"] else "— (not disclosed in free sources)",
        (rnd["announced_at"] or "")[:10] if rnd else "—",
        _fmt_money(val) if val else "— (requires PitchBook)",
        (rnd["lead"] if rnd and rnd["lead"] else "— (not disclosed)"),
        tier1,
        gated(headcount),
        gated(growth),
        f"{c['percentile']:.0f}th pct of {c['cohort_size']} in {c['cohort_key']}{lc}",
        c.get("sector") or "—",
        (c.get("last_signal_at") or "")[:10],
        c.get("human_override") or c.get("recommendation"),
        brief_link,
        (f"{comm['n']} items; latest: {comm_sum['quote'][:80]}…" if comm and comm["n"] and comm_sum
         else "— (no commentary captured yet; X/Blind/podcasts require licenses)"),
    ]


# --------------------------------------------------------------------------
# Tab writers
# --------------------------------------------------------------------------

def write_workbook(verbose: bool = True) -> Path:
    read_human_edits()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    stale_days = thesis()["scoring"]["stale_days"]

    all_scored = scoring.latest_scores(("hot", "watchlist", "pipeline", "stale_review"))

    # 1. Pipeline — every scored, relevant company
    ws = _sheet(wb, "Pipeline", PIPELINE_COLS)
    for c in sorted(all_scored, key=lambda x: -(x["percentile"] or 0)):
        ws.append(_pipeline_row(c))
    _legend(ws)

    # 2. Hot Deals — highest conviction, last 30 days, 60/40 focus split applied
    ws = _sheet(wb, "Hot Deals", PIPELINE_COLS)
    hot = [c for c in all_scored
           if (c.get("human_override") or c["recommendation"]) == "Deep Dive"
           and (scoring._days_since(c["last_signal_at"]) or 999) <= 30]
    for c in scoring.apply_focus_split(hot, limit=15):
        ws.append(_pipeline_row(c))
    _legend(ws)

    # 3. Watchlist
    ws = _sheet(wb, "Watchlist", PIPELINE_COLS)
    for c in [c for c in all_scored if (c.get("human_override") or c["recommendation"]) == "Watch"]:
        ws.append(_pipeline_row(c))
    _legend(ws)

    # 4. Sector of Tomorrow — with the companies sourced INSIDE each cluster (§2b)
    ws = _sheet(wb, "Sector of Tomorrow", [
        "Emerging sector", "Defining terms", "Signal velocity", "Consensus volume",
        "Signal/consensus ratio", "Source diversity", "Talent flow", "Contrarian?",
        "Detected", "Best companies found in this sector", "Evidence (URLs)", "Thesis note"])
    for s in db.q("SELECT * FROM sectors_emerging ORDER BY ratio DESC"):
        ev = json.loads(s["evidence_json"] or "[]")
        found = json.loads(s["companies_json"] or "[]")
        best = "; ".join(
            f"{c['company']} ({c['percentile']:.0f}th pct, {c['recommendation'] or 'unscored'})"
            for c in found) or "— (no company in the pipeline matches this cluster yet)"
        ws.append([s["label"], ", ".join(json.loads(s["terms_json"] or "[]")[:6]),
                   s["signal_velocity"], s["consensus_volume"],
                   round(s["ratio"] or 0, 2), s["source_diversity"],
                   s["talent_flow"],
                   "YES" if s["is_contrarian"] else "", (s["detected_at"] or "")[:10],
                   best,
                   "; ".join(e.get("url", "") for e in ev[:4] if isinstance(e, dict)),
                   (s["thesis_md"] or "")[:200]])
    if ws.max_row == 1:
        ws.append(["— no emerging sector cleared the evidence bar this run (honest empty state;"
                   " widen the lookback window before inventing anything)"] + [""] * 11)
    _legend(ws)

    # 5. Peer Set Activity — investor-by-company events, filterable
    ws = _sheet(wb, "Peer Set Activity", [
        "Investor", "Tier", "Event", "Company / vehicle", "Theme", "Stage",
        "Date", "Thesis shift?", "Deviation", "Source URL"])
    for e in db.q("""SELECT pe.*, i.name inv, i.tier, c.name comp, c.sector, c.stage,
                            s.url, s.payload_json
                     FROM peer_events pe JOIN investors i ON pe.investor_id=i.id
                     LEFT JOIN companies c ON pe.company_id=c.id
                     LEFT JOIN signals s ON pe.source_signal_id=s.id
                     ORDER BY pe.observed_at DESC"""):
        vehicle = e["comp"]
        if not vehicle and e["payload_json"]:
            vehicle = json.loads(e["payload_json"]).get("issuer")
        ws.append([e["inv"], e["tier"], e["event_type"], vehicle or "—",
                   e["sector"] or "—", e["stage"] or "—", (e["observed_at"] or "")[:10],
                   "YES" if e["is_thesis_shift"] else "", e["deviation_score"], e["url"]])
    _legend(ws)

    # 6. Co-investor Heatmap
    ws = _sheet(wb, "Co-investor Heatmap", [
        "Firm A", "Firm B", "Co-investments", "Companies", "Note"])
    pairs = db.q("""
        SELECT i1.name a, i2.name b, COUNT(DISTINCT v1.company_id) n,
               GROUP_CONCAT(DISTINCT c.name) comps
        FROM investments v1 JOIN investments v2
             ON v1.company_id=v2.company_id AND v1.investor_id < v2.investor_id
        JOIN investors i1 ON v1.investor_id=i1.id
        JOIN investors i2 ON v2.investor_id=i2.id
        JOIN companies c ON v1.company_id=c.id AND c.is_synthetic=0
        GROUP BY i1.name, i2.name ORDER BY n DESC LIMIT 100""")
    for p in pairs:
        ws.append([p["a"], p["b"], p["n"], p["comps"], ""])
    if ws.max_row == 1:
        ws.append(["— no co-investment pair observed yet in free sources", "",
                   "", "", "populates as RSS/EDGAR ingest accumulates syndicate data;"
                   " densest with PitchBook/Crunchbase licensed"])
    _legend(ws)

    # 7. News Worth Reading — curated, capped, one-line rationale each
    ws = _sheet(wb, "News Worth Reading", [
        "Title", "Source", "Published", "Why it matters to this fund", "URL"])
    for n in db.q("SELECT * FROM news_items WHERE why_it_matters IS NOT NULL"
                  " ORDER BY relevance_score DESC LIMIT 15"):
        ws.append([n["title"], n["source"], (n["published_at"] or "")[:10],
                   n["why_it_matters"], n["url"]])
    if ws.max_row == 1:
        ws.append(["— curation pass has not run (or nothing met the bar since last digest)",
                   "", "", "", ""])
    _legend(ws)

    # 8. Investor Commentary
    ws = _sheet(wb, "Investor Commentary", [
        "Company", "Platform", "Author", "Credibility", "Sentiment", "Themes",
        "Quote", "Date", "URL"])
    for cm in db.q("""SELECT cm.*, c.name comp FROM commentary cm
                      LEFT JOIN companies c ON cm.company_id=c.id
                      WHERE c.is_synthetic IS NULL OR c.is_synthetic=0
                      ORDER BY cm.observed_at DESC LIMIT 200"""):
        ws.append([cm["comp"] or "(sector-level)", cm["platform"], cm["author"],
                   cm["author_credibility"], cm["sentiment"],
                   ", ".join(json.loads(cm["themes_json"] or "null") or []),
                   cm["quote"], (cm["observed_at"] or "")[:10], cm["url"]])
    if ws.max_row == 1:
        ws.append(["— no commentary captured yet. HN/Reddit are free and populate as pipeline"
                   " companies get discussed; X, Blind, podcasts, Substack threads require"
                   " licenses.", "", "", "", "", "", "", "", ""])
    _legend(ws)

    # 9. Stale — flagged for partner review, never auto-deleted
    ws = _sheet(wb, "Stale", [
        "Company", "Sector", "Stage", "Last signal", "Days quiet",
        "Status", "Action needed"])
    for c in db.q("SELECT * FROM companies WHERE is_synthetic=0 AND last_signal_at IS NOT NULL"
                  " AND julianday('now') - julianday(last_signal_at) > ?"
                  " AND status != 'removed'", (stale_days,)):
        days = int((db.q1("SELECT julianday('now') - julianday(?) d", (c["last_signal_at"],))["d"]))
        ws.append([c["name"], c["sector"] or "—", c["stage"] or "—",
                   (c["last_signal_at"] or "")[:10], days, c["status"],
                   "PARTNER REVIEW — remove or keep? (never auto-deleted)"])
    if ws.max_row == 1:
        ws.append(["— nothing stale: every tracked company has a signal within"
                   f" {stale_days} days", "", "", "", "", "", ""])
    _legend(ws)

    # 10. Provenance — every Pipeline column mapped to its source + freshness
    ws = _sheet(wb, "Provenance", [
        "Pipeline column", "Source", "Fetch mode", "Freshness (latest fetch)",
        "Licence status", "Notes"])
    fresh = {r["name"]: r for r in db.q(
        "SELECT so.name, so.health, so.license_vendor, MAX(s.fetched_at) f,"
        " MAX(s.fetch_mode) fm FROM sources so LEFT JOIN signals s ON s.source_id=so.id"
        " GROUP BY so.name, so.health, so.license_vendor")}

    def prov(col, source_names, note=""):
        fr = [fresh[s] for s in source_names if s in fresh and fresh[s]["f"]]
        latest = max((x["f"] for x in fr), default=None)
        mode = ", ".join(sorted({x["fm"] for x in fr if x["fm"]})) or "—"
        lic = "; ".join(f"requires {fresh[s]['license_vendor']}" for s in source_names
                        if s in fresh and fresh[s]["health"] == "license_required")
        ws.append([col, " + ".join(source_names), mode, latest or "— (no fetch yet)",
                   lic or "free source", note])

    prov("Company / description / sector / stage", ["edgar_formd", "rss_news", "hn"],
         "entity-resolved; sector from deterministic keyword match")
    prov("Last round size / date", ["edgar_formd", "rss_news"],
         "Form D offering amounts + RSS-extracted amounts (regex, no model)")
    prov("Valuation", ["pitchbook"], "null until PitchBook licence")
    prov("Lead investor", ["rss_news"], "regex 'led by …' from real articles")
    prov("Tier 1 count", ["edgar_formd", "rss_news"],
         "observed investments × config tier list — arithmetic, never a model")
    prov("Headcount / 6-month growth", ["coresignal"], "null until Coresignal licence")
    prov("Thesis score", ["edgar_formd", "rss_news", "hn", "github_trending"],
         "computed features -> percentile within (sector,stage) cohort")
    prov("Investor commentary", ["hn", "reddit", "x_gp_watchlist", "blind", "podcasts"],
         "HN/Reddit free; X/Blind/podcasts licence-gated")
    prov("News Worth Reading", ["rss_news", "hn"], "curated with hard caps")
    prov("Sector of Tomorrow", ["arxiv", "hn", "rss_news", "github_trending"],
         "signal-to-consensus clustering + talent flow; companies sourced inside each cluster")
    prov("Customer wins / founder moves", ["rss_news", "hn"],
         "regex-classified from real prose; the matched span is stored as evidence")
    prov("Positioning / customer logos / pricing", ["company_website"],
         "author-written alt text and public pricing pages — never inferred")
    prov("GitHub contributors / commit velocity", ["github_trending"],
         "GitHub API contributors Link header + /stats/participation")
    _legend(ws)

    # 11. Demo Cases — the ONLY tab containing synthetic rows, amber + flagged
    ws = _sheet(wb, "Demo Cases", [
        "Record", "is_synthetic", "Purpose", "Mechanism demonstrated", "Evidence"])
    for c in db.q("SELECT * FROM companies WHERE is_synthetic=1"):
        aliases = db.q("SELECT alias, alias_type, confidence FROM company_aliases"
                       " WHERE company_id=?", (c["id"],))
        if "Stalewatch" in c["name"]:
            purpose = "90-day staleness sweep"
            mech = (f"last_signal_at={c['last_signal_at'][:10]} (backdated 100d) -> appears in"
                    " Stale flow flagged for partner review; NOT deleted")
        else:
            purpose = "entity resolution"
            mech = (f"{len(aliases)} alias variants collapsed into one record: "
                    + "; ".join(f"{a['alias']} ({a['alias_type']}, conf={a['confidence']})"
                                for a in aliases))
        row = [c["name"], "TRUE", purpose, mech, "fictional record — amber = synthetic"]
        ws.append(row)
        for i in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=i).fill = AMBER_FILL
    _legend(ws)

    wb.save(WORKBOOK)
    if verbose:
        print(f"  workbook written: {WORKBOOK} ({len(wb.sheetnames)} tabs)")
    return WORKBOOK


if __name__ == "__main__":
    write_workbook()
