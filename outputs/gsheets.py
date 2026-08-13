"""Live Google Sheets mirror of the workbook, with the same two-way sync contract.

The nine required tabs (plus Provenance and Demo Cases) are pushed to a Google
Sheet partners can open and edit from anywhere. Before every push, the
Recommendation column is read back from the sheet and written to the DB —
**the human value wins** and the disagreement is logged to `partner_actions`,
exactly as with the local .xlsx.

Auth: a Google service account. Without credentials this module does NOT fail the
pipeline — it records `not_configured` in `sheet_sync` and the local workbook
remains the source of truth. Setup is documented in RUNBOOK.md §Google Sheets.

Env:
  GOOGLE_SERVICE_ACCOUNT_JSON  path to the service-account key file
  GSHEET_ID                    id of an existing sheet (preferred), or
  GSHEET_TITLE                 title to create/open by name (default 'Thirdbase Deal Pipeline')
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from engine import db  # noqa: E402
from engine.config import env  # noqa: E402

SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive.file"]

REC_COL_INDEX = 15          # 1-based column O on Pipeline-shaped tabs
EDITABLE_TABS = ("Pipeline", "Hot Deals", "Watchlist")


def credentials_path() -> Path | None:
    p = env("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not p:
        return None
    path = Path(p).expanduser()
    return path if path.exists() else None


def credentials_dict() -> dict | None:
    """The service-account key, from a file OR straight out of the environment.

    A file path is right on a laptop and on Render (a Secret File at
    /etc/secrets/…), but plenty of hosts only offer environment variables, and
    "put this JSON somewhere on disk first" is where a working key quietly turns
    into an unconfigured integration. So the same variable accepts the JSON
    inline, and GOOGLE_SERVICE_ACCOUNT_JSON_B64 accepts it base64-encoded for UIs
    that mangle multi-line values."""
    path = credentials_path()
    if path:
        try:
            return json.loads(path.read_text())
        except (OSError, json.JSONDecodeError):
            return None
    raw = env("GOOGLE_SERVICE_ACCOUNT_JSON")
    if raw and raw.strip().startswith("{"):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return None
    b64 = env("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
    if b64:
        import base64
        try:
            return json.loads(base64.b64decode(b64))
        except Exception:  # noqa: BLE001
            return None
    return None


def configured() -> bool:
    return credentials_dict() is not None


def service_account_email() -> str | None:
    """Who to share the sheet with. This is the single most common reason a
    correctly-configured integration returns 403: the sheet was never shared with
    the robot, and nothing in the Google UI volunteers its address."""
    return (credentials_dict() or {}).get("client_email")


def _project_id() -> str | None:
    return (credentials_dict() or {}).get("project_id")


# Google's failures are precise and its error strings are long; the useful part is
# usually one clause. Each entry turns that clause into the action that fixes it.
def diagnose(detail: str | None) -> dict | None:
    """Turn a Google API error into something a person can act on."""
    if not detail:
        return None
    d = detail.lower()
    # Google names the project in the error text; prefer that over the key file,
    # because the number in the message is the one the console link needs.
    m = re.search(r"project (\d{6,})", detail)
    proj = m.group(1) if m else (_project_id() or "")
    sa = service_account_email() or "the service-account address"
    if "has not been used in project" in d or "is disabled" in d:
        api = ("Google Drive API" if "drive.googleapis.com" in d or "drive api" in d
               else "Google Sheets API" if "sheets" in d else "the required Google API")
        host = "drive.googleapis.com" if "Drive" in api else "sheets.googleapis.com"
        return {"cause": f"{api} is switched off in Google Cloud project"
                         + (f" {proj}" if proj else ""),
                "fix": "Enable BOTH the Google Sheets API and the Google Drive API for that "
                       "project, wait about a minute, then press Test Google Sheet again. "
                       "(Setting GSHEET_ID to an existing sheet avoids needing Drive at all.)",
                "url": f"https://console.developers.google.com/apis/api/{host}/overview"
                       + (f"?project={proj}" if proj else "")}
    if "429" in d or "write requests per minute" in d:
        return {"cause": "Google's per-minute write quota was hit, not a credentials "
                         "problem — the sheet is reachable",
                "fix": "Wait a minute and test again. Each sync now sends the whole "
                       "workbook in two write calls, so this should only appear if "
                       "several syncs overlapped.",
                "url": None}
    if "storagequotaexceeded" in d or "quota" in d and "storage" in d:
        return {"cause": "a service account has no Drive storage of its own, so it "
                         "cannot create a spreadsheet",
                "fix": f"Create the sheet yourself in your own Drive, share it with {sa} "
                       "as an Editor, and set GSHEET_ID to the id in its URL.",
                "url": "https://sheets.new"}
    if "permissiondenied" in d or "does not have permission" in d or "[403]" in d:
        return {"cause": "the service account cannot open that spreadsheet",
                "fix": f"Open the sheet, press Share, and give {sa} Editor access.",
                "url": None}
    if "[404]" in d or "requested entity was not found" in d:
        return {"cause": "GSHEET_ID does not point at a spreadsheet the service account "
                         "can see",
                "fix": "Copy the id out of the sheet URL between /d/ and /edit, and make "
                       f"sure the sheet is shared with {sa}.",
                "url": None}
    return None


def status() -> dict:
    """Configured is not the same as working. Reporting only the first would have
    shown this integration as connected while every sync had been failing for a
    day — the credentials were found, and Google was refusing them."""
    last = db.q1("SELECT * FROM sheet_sync ORDER BY synced_at DESC LIMIT 1")
    last_d = dict(last) if last else None
    detail = (last_d or {}).get("detail") if (last_d or {}).get("status") == "error" else None
    return {"configured": configured(),
            "credentials": str(credentials_path()) if credentials_path()
                           else ("inline env JSON" if configured() else None),
            "service_account_email": service_account_email(),
            "sheet_id_set": bool(env("GSHEET_ID")),
            "ok": bool(last_d and last_d.get("status") == "ok"),
            "url": (last_d or {}).get("spreadsheet_url"),
            "reason": None if configured() else
                      "GOOGLE_SERVICE_ACCOUNT_JSON not set (path or inline JSON)",
            "last_error": detail,
            "diagnosis": diagnose(detail),
            "last_sync": last_d}


def _client():
    import gspread
    from google.oauth2.service_account import Credentials
    creds = Credentials.from_service_account_info(credentials_dict(), scopes=SCOPES)
    return gspread.authorize(creds), creds


def _open_sheet(gc, creds):
    """Open by id when we have one — that path uses ONLY the Sheets API.

    Opening or creating by title goes through Drive, which is a second API to
    enable and, for `create`, a storage quota a service account does not have.
    Both failures look identical from the outside (a 403 on sync), so the id path
    is preferred and the fallback says out loud what it is about to depend on."""
    sheet_id = env("GSHEET_ID")
    if sheet_id:
        return gc.open_by_key(sheet_id.strip())
    title = env("GSHEET_TITLE", "Thirdbase Deal Pipeline")
    try:
        return gc.open(title)          # Drive API: search by name
    except Exception:  # noqa: BLE001 — not found, create it
        sh = gc.create(title)          # Drive API + service-account storage
        share_to = env("GSHEET_SHARE_WITH")
        if share_to:
            for addr in share_to.split(","):
                try:
                    sh.share(addr.strip(), perm_type="user", role="writer")
                except Exception:  # noqa: BLE001
                    pass
        print(f"  created Google Sheet '{title}' — {sh.url}")
        print("  NOTE: a service-account sheet is not in your Drive UI until shared."
              " Set GSHEET_SHARE_WITH=you@gmail.com, or use GSHEET_ID of a sheet you"
              " already shared with the service-account email.")
        return sh


def _retry(call, attempts: int = 4):
    """Google's write quota is per minute, so a 429 is a 'wait', not a 'no'.

    Batching means we should rarely see one; when we do — a partner pressing Test
    twice, two runs overlapping — sleeping and retrying is the correct response,
    and failing the whole sync while telling the user their credentials are broken
    is not."""
    import time
    for i in range(attempts):
        try:
            return call()
        except Exception as exc:  # noqa: BLE001
            msg = str(exc)
            transient = "429" in msg or "Quota exceeded" in msg or "[503]" in msg
            if not transient or i == attempts - 1:
                raise
            time.sleep(min(2 ** i * 5, 40))     # 5s, 10s, 20s — quota is per minute
    return None


def _header_requests(sheet_id: int, n_cols: int) -> list[dict]:
    """Freeze + style row 1, as raw API requests so they can ride in one batch."""
    return [
        {"updateSheetProperties": {
            "properties": {"sheetId": sheet_id,
                           "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount"}},
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                      "startColumnIndex": 0, "endColumnIndex": max(n_cols, 1)},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True,
                               "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "backgroundColor": {"red": 0.12, "green": 0.23, "blue": 0.34}}},
            "fields": "userEnteredFormat(textFormat,backgroundColor)"}},
    ]


def pull_human_edits(sh, partner: str = "partner", verbose: bool = True) -> int:
    """Read Recommendation edits from the sheet into the DB. Human value wins."""
    n = 0
    seen: set[str] = set()
    for tab in EDITABLE_TABS:
        try:
            ws = sh.worksheet(tab)
        except Exception:  # noqa: BLE001 — tab not created yet
            continue
        rows = ws.get_all_values()
        for row in rows[1:]:
            if not row or not row[0] or row[0].startswith("Legend:"):
                continue
            name = row[0]
            if name in seen or len(row) < REC_COL_INDEX:
                continue
            seen.add(name)
            rec_val = (row[REC_COL_INDEX - 1] or "").strip()
            if rec_val not in ("Pass", "Watch", "Deep Dive"):
                continue
            comp = db.q1("SELECT id FROM companies WHERE name=? AND is_synthetic=0", (name,))
            if not comp:
                continue
            score = db.q1("SELECT id, recommendation, human_override, composite, features_json"
                          " FROM scores WHERE company_id=? ORDER BY scored_at DESC, id DESC"
                          " LIMIT 1", (comp["id"],))
            if not score:
                continue
            effective = score["human_override"] or score["recommendation"]
            if rec_val != effective:
                db.execute("UPDATE scores SET human_override=? WHERE id=?",
                           (rec_val, score["id"]))
                db.insert("partner_actions", {
                    "company_id": comp["id"], "partner": partner, "action": "override",
                    "score_at_time": score["composite"],
                    "features_at_time_json": score["features_json"],
                    "note": f"google sheet edit: {effective} -> {rec_val} ({tab})",
                    "created_at": db.now_iso()})
                n += 1
    if verbose and n:
        print(f"  google sheet: {n} human edit(s) pulled into DB (human value wins)")
    return n


def _tab_data(xlsx_path: Path) -> list[tuple[str, list[list]]]:
    """Read the generated workbook so the Sheet is a faithful mirror — one
    renderer, two destinations, no second source of truth."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        out.append((name, rows))
    return out


def sync(xlsx_path: Path | None = None, verbose: bool = True) -> dict:
    from engine.config import OUTPUT_DIR
    xlsx_path = xlsx_path or (OUTPUT_DIR / "deal_pipeline.xlsx")

    if not configured():
        rec = {"status": "not_configured", "tabs_written": 0, "edits_pulled": 0,
               "detail": "GOOGLE_SERVICE_ACCOUNT_JSON not set — local workbook"
                         " remains source of truth", "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  google sheets: skipped — {rec['detail']}")
        return rec
    if not xlsx_path.exists():
        # Same ephemeral-disk trap as the download button: the rows are in the
        # database, only the artefact is missing. Build it rather than refuse.
        try:
            from outputs.excel import ensure_workbook
            xlsx_path = ensure_workbook()
        except Exception as e:  # noqa: BLE001
            rec = {"status": "error", "detail": f"workbook could not be built: {e}"[:400],
                   "tabs_written": 0, "edits_pulled": 0, "synced_at": db.now_iso()}
            db.insert("sheet_sync", rec)
            return rec

    try:
        gc, creds = _client()
        sh = _open_sheet(gc, creds)
        edits = pull_human_edits(sh, verbose=verbose)   # read BEFORE writing
        tabs = _tab_data(xlsx_path)
        # Eleven tabs used to cost five write calls each — clear, resize, update,
        # freeze, format — which is fifty-five requests against a sixty-per-minute
        # quota, for a sync that also took over a minute of round trips. It worked
        # exactly until it didn't:
        #   [429] Quota exceeded for 'Write requests per minute per user'
        # A rate limit hit on the FIRST successful sync is not a Google problem,
        # it is a design problem: the whole workbook is one payload and should be
        # sent as one. Batched, a refresh is two write calls regardless of how
        # many tabs the workbook grows to.
        existing = {w.title: w for w in sh.worksheets()}
        created: list[tuple[str, int]] = []
        for name, rows in tabs:
            n_rows = max(len(rows), 2)
            n_cols = max((len(r) for r in rows), default=1)
            ws = existing.get(name)
            if ws is None:
                ws = _retry(lambda: sh.add_worksheet(title=name, rows=n_rows, cols=n_cols))
                existing[name] = ws
                created.append((name, n_cols))
            elif ws.row_count < n_rows or ws.col_count < n_cols:
                # only when the data outgrew the grid — a resize on every sync is
                # a write call spent to change nothing
                _retry(lambda: ws.resize(rows=max(n_rows, ws.row_count),
                                         cols=max(n_cols, ws.col_count)))

        # 1 call: wipe every tab's old contents
        _retry(lambda: sh.values_batch_clear(
            {"ranges": [f"'{name}'" for name, _ in tabs]}))
        # 1 call: write every tab's new contents
        _retry(lambda: sh.values_batch_update({
            "valueInputOption": "RAW",
            "data": [{"range": f"'{name}'!A1", "values": rows}
                     for name, rows in tabs if rows]}))

        # Header styling is a property of the tab, not of the data, so it is
        # applied once when the tab is born rather than re-sent on every refresh.
        if created:
            _retry(lambda: sh.batch_update({"requests": [
                r for name, n_cols in created
                for r in _header_requests(existing[name].id, n_cols)]}))
        for stale in set(existing) - {n for n, _ in tabs} - {"Sheet1"}:
            pass  # never delete partner-created tabs
        rec = {"status": "ok", "spreadsheet_id": sh.id, "spreadsheet_url": sh.url,
               "tabs_written": len(tabs), "edits_pulled": edits,
               "detail": None, "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  google sheets: {len(tabs)} tabs synced -> {sh.url}")
        return rec
    except Exception as exc:  # noqa: BLE001 — never kill the pipeline for a sync
        rec = {"status": "error", "tabs_written": 0, "edits_pulled": 0,
               "detail": f"{type(exc).__name__}: {exc}"[:400], "synced_at": db.now_iso()}
        db.insert("sheet_sync", rec)
        if verbose:
            print(f"  ! google sheets sync failed — {rec['detail']}")
            # The raw Google error names a symptom; this names the action. Without
            # it the operator reads "[403] APIError" and has nowhere to go.
            d = diagnose(rec["detail"])
            if d:
                print(f"    cause: {d['cause']}")
                print(f"    fix:   {d['fix']}")
                if d.get("url"):
                    print(f"    open:  {d['url']}")
        return rec


if __name__ == "__main__":
    if "--status" in sys.argv:
        print(json.dumps(status(), indent=2, default=str))
    else:
        print(json.dumps(sync(), indent=2, default=str))
